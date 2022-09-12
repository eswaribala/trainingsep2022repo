# -*- coding: utf-8 -*-
"""
Created on Mon Jun 11 15:02:09 2018

@author: Balasubramaniam
"""

from openpyxl import load_workbook
filePath="F:/citi_ml_jun2018/day1/GDP.xlsx"
wb=load_workbook(filePath,read_only=True,data_only=True)
sheetRef=wb.get_sheet_by_name("GDP")
for row in range(6,70):
    print(sheetRef.cell(column=1,row=row).value,end="\t")

