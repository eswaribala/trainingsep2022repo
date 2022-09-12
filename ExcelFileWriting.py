# -*- coding: utf-8 -*-
"""
Created on Mon Jun 11 15:15:33 2018

@author: Balasubramaniam
"""
from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill, Font,Fill,Border
from openpyxl.styles.colors import RED

filePath="F:/citi_ml_jun2018/day1/AnnualReport.xlsx"
wb=load_workbook(filePath,read_only=False)
#import calendar
#
#for month in calendar.month_name:
#   if not(len(month)==0): 
#      print(month)
#      wb.create_sheet(month+"_2018")
#wb.save(filePath)
import datetime
currentMonth=datetime.date.today().strftime("%B")
print(currentMonth)
sheetRef=wb.get_sheet_by_name(currentMonth+"_2018")
import random
for row in range(1,100):
    for col in range(1,10):
        sheetRef.cell(column=col,row=row,value="%d" % (random.randint(1,10000)))
        
wb.save(filePath)
for row in range(1,100):
  for col in range(1,10):
      cell=sheetRef.cell(row=row,column=col)
      if(int(cell.value) % 10 == 0):       
           cell.font = Font(size=18,color=RED)
wb.save(filePath)
wb.close()

